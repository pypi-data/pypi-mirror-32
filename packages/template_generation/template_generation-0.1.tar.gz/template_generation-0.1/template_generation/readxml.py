import os
from xml.dom import minidom

from openpyxl import Workbook

import xlrd
import csv

def get_path_from_current(*path):
    return os.path.join(os.path.dirname(__file__), *path)


class ReadXml:

    def __init__(self, sourcePath, targetPath):
        self.sourcePath = sourcePath
        self.targetPath = targetPath

    def get_file_list(self):
        fileList = []
        for root, dirs, files in os.walk(self.sourcePath):
            for filename in files:
                fileList.append(filename)
        return fileList

    def read_write_data(self):
        fileList = self.get_file_list()
        print(fileList)
        for filename in fileList:
            xmlFile = get_path_from_current(self.sourcePath, filename)
            print(xmlFile)
            xmldoc = minidom.parse(xmlFile)
            itemList = xmldoc.getElementsByTagName('Worksheet')
            sheetList = []
            fieldFullList = []
            textFullList = []
            dataList = []
            lineList = []
            positionList = []
            sourceIdList = []
            i = 1
            for item in itemList:
                if i > 2:
                    sheet = item.attributes['ss:Name'].value
                    if item.hasChildNodes():
                        tables = item.getElementsByTagName("Table")
                        for table in tables:
                            textList = []
                            fieldList = []
                            headerList = []
                            subLineList = []
                            subSource = []
                            rows = table.getElementsByTagName("Row")
                            j = 1
                            field = ''
                            text = ''
                            for row in rows:
                                cells = row.getElementsByTagName("Cell")
                                if j == 5:
                                    for cell in cells:
                                        data = cell.getElementsByTagName("Data")
                                        for d in data:
                                            for childNode in d.childNodes:
                                                field = childNode.nodeValue
                                                fieldList.append(field)
                                if j == 7:
                                    for cell in cells:
                                        if cell.hasAttribute('ss:MergeAcross'):
                                            merge = int(cell.attributes['ss:MergeAcross'].value) + 1
                                        else:
                                            merge = 1
                                        data = cell.getElementsByTagName("Data")
                                        for d in data:
                                            for childNode in d.childNodes:
                                                header = childNode.nodeValue
                                                m = 1
                                                while m <= merge:
                                                    headerList.append(header)
                                                    m += 1
                                if j == 8:
                                    for cell in cells:
                                        data = cell.getElementsByTagName("Data")
                                        for d in data:
                                            for childNode in d.childNodes:
                                                ltext = childNode.nodeValue
                                                pos = ltext.index('\n')
                                                text = ltext[:pos].rstrip('*')
                                                textList.append(text)

                                if j >= 9:
                                    line = j - 8
                                    position = 1
                                    sourceId = ''

                                    for cell in cells:
                                        data = cell.getElementsByTagName("Data")
                                        for d in data:
                                            for childNode in d.childNodes:
                                                v = childNode.nodeValue
                                                sheetList.append(sheet)
                                                fieldFullList.append(fieldList[position - 1])
                                                textFullList.append(textList[position - 1])
                                                dataList.append(v)
                                                lineList.append(line)
                                                subLineList.append(line)
                                                positionList.append(position)
                                                if headerList[position - 1] == 'Key':
                                                    sourceId += '/' + v if sourceId != '' else v

                                        position += 1
                                    subSource.append(sourceId)
                                j += 1
                        print(subSource)
                        print(subLineList)
                        for k in subLineList:
                            sourceIdList.append(subSource[int(k - 1)])
                i += 1
            print(lineList)
            print(sourceIdList)
            wb = Workbook()
            ws = wb.active
            # ws.title = "New Title"
            excel_header = ["FileName", "SheetName", "LineNo", "SourceID", "FieldPosition", "FieldName", "FieldText",
                            "SourceValue", "TargetValue"]
            ws.append(excel_header)
            for s in range(0, len(sheetList)):
                ws.cell(row=s + 2, column=1, value=filename)
                ws.cell(row=s + 2, column=2, value=sheetList[s])

            for l in range(0, len(lineList)):
                ws.cell(row=l + 2, column=3, value=lineList[l])

            for s in range(0, len(sourceIdList)):
                ws.cell(row=s + 2, column=4, value=sourceIdList[s])

            for p in range(0, len(positionList)):
                ws.cell(row=p + 2, column=5, value=positionList[p])

            for f in range(0, len(fieldFullList)):
                ws.cell(row=f + 2, column=6, value=fieldFullList[f])

            for t in range(0, len(textFullList)):
                ws.cell(row=t + 2, column=7, value=textFullList[t])

            for d in range(0, len(dataList)):
                ws.cell(row=d + 2, column=8, value=dataList[d])
            target = get_path_from_current(self.targetPath, filename.rstrip('.xml') + '.xlsx')
            wb.save(target)


if __name__ == '__main__':
    ReadXml("source", "excel").read_write_data()
