__author__ = 'I071034'
import os
from xml.dom import minidom
from xml.etree import ElementTree
from openpyxl import Workbook
import traceback
import openpyxl
import xlrd
import csv
import logging

logger = logging.getLogger(__name__)


def get_path_from_current( *path):
    return os.path.join(os.path.dirname(__file__), *path)


def get_file_list( path):
    fileList = []
    for root, dirs, files in os.walk(path):
        for filename in files:
            fileList.append(filename)
    return fileList

class FormatError(Exception):
    def __init__(self, message):
        self.strerror = message
    def __str__(self):
        return repr(self.strerror)

class ProcessFile:

    def __init__(self, os_path, xmlPath, excelPath, csvPath):
        self.os_path = os_path
        self.xmlPath = xmlPath
        self.excelPath = excelPath
        self.csvPath = csvPath

    class ProcessFileDecorator:
        def process_file_exception(func):
            def wrapper(*args, **kwargs):
                try:
                    rv = func(*args, **kwargs)
                    return rv
                except FormatError as format_error:
                    print("The file format is invalid: {0}".format(format_error.strerror))
                    logging.error(format_error)
                    raise
                except IOError as io_error:
                    print("The file could not be read: {0}".format(io_error.strerror))
                    logging.error(io_error)
                    raise

            return wrapper


    @ProcessFileDecorator.process_file_exception
    def xml_to_excel(self):
        fileList = get_file_list(get_path_from_current(self.os_path, self.xmlPath))
        for filename in fileList:
            flie_name = filename.rstrip('.xml')
            xmlFile = get_path_from_current(self.os_path, self.xmlPath, filename)
            xmldoc = minidom.parse(xmlFile)
            itemList = xmldoc.getElementsByTagName('Worksheet')
            sheetList = []

            i = 1
            for item in itemList:
                if i > 2:
                    sheet = item.attributes['ss:Name'].value
                    sheetList.append(sheet)
                    fieldFullList = []
                    textFullList = []
                    sourceIdList = []
                    dataList = []
                    positionList = []
                    lineList = []
                    if item.hasChildNodes():
                        tables = item.getElementsByTagName("Table")
                        for table in tables:
                            textList = []
                            fieldList = []
                            headerList = []
                            subSource = []
                            rows = table.getElementsByTagName("Row")
                            j = 1
                            field = ''
                            text = ''
                            for row in rows:
                                cells = row.getElementsByTagName("Cell")
                                if j == 5:
                                    for cell in cells:
                                        data = cell.getElementsByTagName("Data")
                                        for d in data:
                                            for childNode in d.childNodes:
                                                field = childNode.nodeValue
                                                fieldList.append(field)
                                if j == 7:
                                    for cell in cells:
                                        if cell.hasAttribute('ss:MergeAcross'):
                                            merge = int(cell.attributes['ss:MergeAcross'].value) + 1
                                        else:
                                            merge = 1
                                        data = cell.getElementsByTagName("Data")
                                        for d in data:
                                            for childNode in d.childNodes:
                                                header = childNode.nodeValue
                                                m = 1
                                                while m <= merge:
                                                    headerList.append(header)
                                                    m += 1
                                if j == 8:
                                    for cell in cells:
                                        data = cell.getElementsByTagName("Data")
                                        for d in data:
                                            for childNode in d.childNodes:
                                                ltext = childNode.nodeValue
                                                pos = ltext.index('\n')
                                                text = ltext[:pos].rstrip('*')
                                                textList.append(text)

                                if j >= 9:
                                    line = j - 8
                                    position = 1
                                    sourceId = ''

                                    for cell in cells:
                                        data = cell.getElementsByTagName("Data")
                                        for d in data:
                                            for childNode in d.childNodes:
                                                v = childNode.nodeValue
                                                fieldFullList.append(fieldList[position - 1])
                                                textFullList.append(textList[position - 1])
                                                dataList.append(v)

                                                lineList.append(line)
                                                positionList.append(position)
                                                if headerList[position - 1] == 'Key':
                                                    sourceId += '/' + v if sourceId != '' else v

                                        position += 1
                                    subSource.append(sourceId)
                                j += 1

                        for k in lineList:
                            sourceIdList.append(subSource[int(k - 1)])

                        wb = Workbook()
                        ws = wb.active
                        # ws.title = "New Title"
                        excel_header = ["LineNo", "SourceID","FieldPosition", "FieldName", "FieldText",
                                        "TargetValue"]
                        ws.append(excel_header)
                        for l in range(0, len(lineList)):
                            ws.cell(row=l + 2, column=1, value=lineList[l])

                        for s in range(0, len(sourceIdList)):
                            ws.cell(row=s + 2, column=2, value=sourceIdList[s])

                        for p in range(0, len(positionList)):
                            ws.cell(row=p + 2, column=3, value=positionList[p])

                        for f in range(0, len(fieldFullList)):
                            ws.cell(row=f + 2, column=4, value=fieldFullList[f])

                        for t in range(0, len(textFullList)):
                            ws.cell(row=t + 2, column=5, value=textFullList[t])

                        for d in range(0, len(dataList)):
                            ws.cell(row=d + 2, column=6, value=dataList[d])
                        target = get_path_from_current(self.os_path, self.excelPath, "%s_%s.xlsx" % (flie_name, sheet))
                        wb.save(target)
                i += 1

    @ProcessFileDecorator.process_file_exception
    def excel_to_csv(self):
        file_list = get_file_list(get_path_from_current(self.os_path, self.excelPath))
        for file in file_list:
            excel_file = get_path_from_current(self.os_path, self.excelPath, file)
            with xlrd.open_workbook(excel_file) as wb:
                sh = wb.sheet_by_index(0)  # or wb.sheet_by_name('name_of_the_sheet_here')
                csv_file = get_path_from_current(self.os_path, self.csvPath,file.rstrip(".xlsx") + ".csv")
                with open(csv_file, 'w', newline="") as f:
                    c = csv.writer(f)
                    for r in range(sh.nrows):
                        c.writerow(sh.row_values(r))

    @ProcessFileDecorator.process_file_exception
    def xml_to_csv(self):
        fileList = get_file_list(get_path_from_current(self.os_path, self.xmlPath))
        for filename in fileList:
            flie_name = filename.rstrip('.xml')
            xmlFile = get_path_from_current(self.os_path, self.xmlPath, filename)
            xmldoc = minidom.parse(xmlFile)

            title = xmldoc.getElementsByTagName('Title')
            for t in title:
                for childNode in t.childNodes:
                    obj_name = childNode.nodeValue

            itemList = xmldoc.getElementsByTagName('Worksheet')
            i = 1
            for item in itemList:
                if i > 2:
                    sheet_name = item.attributes['ss:Name'].value
                    file_name = obj_name + '_' + sheet_name + '.csv'
                    csv_file = get_path_from_current(self.os_path, self.csvPath, file_name)
                    with open(csv_file, 'w', newline="", encoding='utf-8') as f:
                        c = csv.writer(f)
                        if item.hasChildNodes():
                            tables = item.getElementsByTagName("Table")
                            for table in tables:

                                rows = table.getElementsByTagName("Row")
                                j = 1
                                for row in rows:
                                    cells = row.getElementsByTagName("Cell")
                                    if j == 5:
                                        field_list = []
                                        for cell in cells:
                                            data = cell.getElementsByTagName("Data")
                                            for d in data:
                                                for childNode in d.childNodes:
                                                    field = childNode.nodeValue
                                                    field_list.append(field)
                                        c.writerow(field_list)

                                    if j >= 9:
                                        data_list = []
                                        k = 1
                                        for cell in cells:
                                            if cell.hasAttribute('ss:Index'):
                                                index = int(cell.attributes['ss:Index'].value)
                                                bal = index -k
                                                for l in range(bal):
                                                    data_list.append('')
                                                    k += 1
                                            data = cell.getElementsByTagName("Data")
                                            for d in data:
                                                for childNode in d.childNodes:
                                                    v = childNode.nodeValue
                                                    data_list.append(v)
                                            k += 1
                                        if len(data_list) < len(field_list):
                                            for p in range(len(field_list)-len(data_list)):
                                                data_list.append('')
                                        c.writerow(data_list)
                                    j += 1

                i += 1
