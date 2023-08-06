#!/usr/bin/python3

#-----------------------------
VERSION="1.0"
DATETIME="2014-06-25 at 18:21:47"
MODULE="SysGen"
NAME="SysGen"
DESCRIPTION="A FPGA system generator"
#-----------------------------

import os, sys, logging, argparse, subprocess
try:
	import configparser
except ImportError:
	import configparser as configparser
	
from openpyxl import Workbook
import datetime, tempfile
#======================LOGGING CONFIGURATION===========================
sys.path.append(os.path.normpath(os.path.join(os.path.dirname(__file__), "..")))
from Utilities import ConsoleInterface
ConsoleInterface.ConfigLogging(Version="1.0", ModuleName="SysGen")
#======================================================================
from Utilities import ColoredLogging, Timer, Misc, DataMining
from SysGen.HW.HwModel import HwModel
#import SysGen
from SysGen import XmlLibManager, Module
import SysGen
import pkgutil

#=======================================================================
def ShowGUI(Options):
	"""
	Initialize and launch GUI.
	"""
	if Options.lib is None: Library=SysGen.BBLIBRARYPATH
	else: Library=Options.lib
	logging.error("No GUI available yet.")

#=======================================================================
def ListModules(Options):
	"""
	Display a list of modules found in the library.
	"""
	Library=XmlLibManager.XmlLibManager(SysGen.BBLIBRARYPATH)
	for Mod in Library.ListModules():
		print(Mod.Name)
	return
	
	
#=======================================================================
def EvalModuleCmd(Options):
	"""
	Implement specified module on target architecture and fetch FPGA resource usage.
	"""
	# Fetch module in library
	Library=XmlLibManager.XmlLibManager(SysGen.BBLIBRARYPATH)
	OriMod=Library.Module(Options.module, [Library,])
	Mod=OriMod.Copy()
	if Mod is None: 
		logging.error("No such module '{0}' in library.".format(Options.module))
		return None
	Archi    = Options.architecture
	ParamVal = {}
	for PAssignment in Options.parameter:
		if PAssignment=='': continue
		try: PName, Value=PAssignment.split('=', 2)
		except: 
			logging.error("Cannot recognize expression '{0}'. Expected format: <Name>=<Value>.".format(PAssignment))
			return None
#		for PNameValue in PNameValues:
#			PName, Value=PNameValue.split(':')
		ParamVal[PName]=Value
	RemoteHost = Options.host
	TargetHwModel=HwModel(Archi)
	HWRsc=TargetHwModel.ModuleResources(
				Mod      = Mod, 
				Library=Library,
				ParamVal = ParamVal, 
				RemoteHost=RemoteHost
				)
	if HWRsc is None:
		logging.error("Failed to evaluate resources.")
		return None
	else:
		logging.info("Used resources : {0}".format(HWRsc))
		# Update XML
		OriMod.UpdateResources(FpgaId=TargetHwModel.GetFPGA(FullId=False), HWRsc=HWRsc)
		OriMod.DumpToLibrary() 
		return HWRsc
	
#=======================================================================
#def EvalModule(Mod, Archi, ParamVal={}, Library=None, RemoteHost=None):
#	"""
#	Implement specified module on target architecture and fetch FPGA resource usage.
#	"""
#	# Fetch architecture in library
#	TargetHW=HwModel(Archi)
#	if not TargetHW.IsSupported():
#		logging.error("Architecture {0} not supported. {0}'s configuration must be provided in library.".format(Archi))
#		return None
#	
#	logging.info("Start evaluation of FPGA resources usage for module '{0}' on architecture '{1}' with parameters '{2}'".format(Mod, TargetHW, ParamVal))
#	
#	for PName, Value in ParamVal.items():
#		if PName in Mod.Params:
#			logging.info("\t> Set '{0}' to {1}.".format(PName, Value))
#			Mod.EditParam(PName, Value)
#		else:
#			logging.error("No such parameter '{0}' in module '{1}'.".format(PName, Mod.Name))
#	Mod.Reload()
#	if Library: 
#		Mod.IdentifyServices(Library.Services)
#		
#	OutputPath=tempfile.mkdtemp(suffix='', prefix='RscEval_{0}_{1}_'.format(Mod.Name, Timer.TimeStamp()), dir=None)
#	SrcPath=os.path.join(OutputPath, "src")
##	SynPath=os.path.join(OutputPath, "synthesis")
#	os.makedirs(SrcPath)
##	os.makedirs(SynPath)
##	WrappedMod=Module.GenerateEmptyWrapper(Mod=Mod, CopyMod=False)
#	WrappedMod=Mod.Copy()
#	WrappedMod.GenSrc( 
#		Synthesizable = True, 
#		OutputDir     = SrcPath, 
#		TestBench     = False, 
#		IsTop         = True,
#		ModInstance   = None,
#		NewPkg        = WrappedMod.IsAbstract(),
##		TBValues      = TBValues,
##		TBClkCycles   = TBClkCycles,
##		CommunicationService=CommunicationService
#		Recursive=True,
#		)
#	
#	HWRsc=TargetHW.ModuleResources(Mod=WrappedMod, Library=Library, OutputPath=OutputPath, RemoteHost=RemoteHost)
#	if HWRsc is None:
#		logging.error("Failed to evaluate resources.")
#		Misc.CleanTempFiles(OutputPath)
#		return None
#	else:
#		logging.info("Used resources : {0}".format(HWRsc))
#		# Update XML
#		Mod.UpdateResources(FpgaId=TargetHW.GetFPGA(FullId=False), HWRsc=HWRsc)
#		Mod.DumpToLibrary() 
#		Misc.CleanTempFiles(OutputPath)
#		return HWRsc
	
#=======================================================================
def ModelNoCResourcesCmd(Options):
	"""
	Implement specified module on target architecture and fetch FPGA resource usage.
	"""
	Archi      = Options.architecture
	OutputPath = Options.output
	RemoteHost     = Options.host
	return ModelNoCResources(Archi=Archi, OutputPath=OutputPath, RemoteHost=RemoteHost)
	
#=======================================================================
def ModelNoCResources(Archi, OutputPath=None, RemoteHost=None):
	"""
	Estimate resource for specified Module for each combination of specified parameters.
	And perform data mining to generate a model.
	"""
	if OutputPath is None: 
		OutputPath=tempfile.mkdtemp(prefix='NoCRscModels_{0}_{1}'.format(Archi, Timer.TimeStamp()), dir='./')
	else:
		OutputPath=os.path.abspath(OutputPath)
		if not os.path.isdir(OutputPath): os.makedirs(OutputPath)
		
	RscIniPath=os.path.join(OutputPath, "CollectedResources_{0}.ini".format(Timer.TimeStamp()))
	Config=configparser.RawConfigParser()
	
	TargetHwModel=HwModel(Archi)
	
	ModelValues={'LUT':[], 'REGISTER':[], 'RAM':[]}
	
	for FifoDepth in [96,]:#2, 4, 16, 32,]:# 
		for FlitWidth in [64,]:#8, 16, 32, 64,]: # 
			for DimX, DimY in [(5,5), (9,9), (12,12)]:
				
				# Fetch module in library (reload library for each iteration)
				Library=XmlLibManager.XmlLibManager(SysGen.BBLIBRARYPATH)
				Mod=Library.Module("AdOCNet", [Library,])
				if Mod is None: 
					logging.error("No such module '{0}' in library.".format("AdOCNet"))
					return None
					
				Rsc=TargetHwModel.ModuleResources(
							Mod      = Mod, 
							Library=Library,
							ParamVal = {'FlitWidth':FlitWidth, 'InputFifoDepth_Table': [FifoDepth for i in range(DimX*DimY)], 'DimX':DimX, 'DimY':DimY, 'NbInputs_Table': [5 for i in range(DimX*DimY)], 'NbOutputs_Table': [5 for i in range(DimX*DimY)],}, 
							RemoteHost=RemoteHost
							)
				if Rsc is None:
					logging.error("[ModelNoCResources] Failed to evaluate resources for module '{0}'.".format(Mod))
					RscDict={}
					return None
				else:
					RscDict=Rsc.GetResourcesDict()
#				RscDict={'LUT':(5*DimX,-1,-1), 'REGISTER':(3*DimX,-1,-1), 'RAM':(1*DimX,-1,-1)}
				
				ConfigName="{0}_{1}*{2}_Flit{3}_Buffer{4}".format('AdOCNet', DimX, DimY, FlitWidth, FifoDepth)
				Config.add_section(ConfigName)
				Config.set(ConfigName, 'DimX', DimX)
				Config.set(ConfigName, 'DimY', DimY)
				Config.set(ConfigName, 'DimX*DimY', DimX*DimY)
				Config.set(ConfigName, 'FlitWidth', FlitWidth)
				Config.set(ConfigName, 'FifoDepth', FifoDepth)
				for Item in ('LUT', 'REGISTER', 'RAM'):
					if Item in RscDict:
						Config.set(ConfigName, Item, RscDict[Item][0])
						ModelValues[Item].append( (DimX*DimY, RscDict[Item][0]) )
					else:
						logging.warning("[ModelNoCResources] No such key '{0}' in resource dictionary {1}.".format(Item, RscDict))
					
				with open(RscIniPath, 'w+') as ConfFile:
					Config.write(ConfFile)
	logging.info("INI Resource file generated: '{0}'".format(RscIniPath))
	XlsPath=RscIniToXls(RscIniPath, HwArchiName=Archi, OutputPath=OutputPath, ParamOrder=('DimX', 'DimY', 'DimX*DimY', 'FlitWidth', 'FifoDepth', 'lut', 'register', 'ram'))
	
	
	logging.debug("Model values: '{0}'".format(ModelValues))
		
	LinearExpressions=DataMining.LinearRegression(ModelValues, Variable='DimX*DimY')
	for RscType, LinearExpression in LinearExpressions.items():
		logging.info(" * Model for '{0}': {1}".format(RscType, LinearExpression))
		
	logging.info("XLS Resource file generated: '{0}'".format(XlsPath))
	
	Answer=''
	while not Answer in ('Y', 'N'):
		Answer=input("Would you like to open the generated XLS file? Y/N\n").upper()
	if Answer=='Y':
		if sys.platform.startswith('linux'):
			subprocess.call(["xdg-open", XlsPath])
		else:
			os.startfile(XlsPath)
	
	return XlsPath
	
#=======================================================================
def RscIniToXls(RscIniPath, HwArchiName, OutputPath, ParamOrder):
	"""
	Generate XLS from ini resource file.
	"""
	if OutputPath is None: 
		OutputPath=tempfile.mkdtemp(prefix='RscIniToXls_{0}_'.format(Timer.TimeStamp()))
	else:
		OutputPath=os.path.abspath(OutputPath)
		if not os.path.isdir(OutputPath): os.makedirs(OutputPath)
		
	wb = Workbook()

	# grab the active worksheet
	ws = wb.active

	Config=configparser.RawConfigParser()
	Config.read(RscIniPath)
	
	ws['A1'] = "Resources collection date:" # Data can be assigned directly to cells
	ws['A2'] = datetime.datetime.now() # Python types will automatically be converted
	
	ws.append(["",]) # Rows can be appended
	ws.append(["Target:",])
	ws.append([HwArchiName,])
	
	ws.append(["",])
	ws.append(["",]+list(ParamOrder)) 
	for Row, Section in enumerate(Config.sections()):
		ws.append([Section,]+[Config.getint(Section, Item) if Config.has_option(Section, Item) else '' for Item in ParamOrder])
		
	XlsPath=os.path.join(OutputPath, "CollectedResources_{0}.xlsx".format(Timer.TimeStamp()))
	# Save the file
	wb.save(XlsPath)
	
	return XlsPath
	
#=======================================================================
def ParseOptions():
	"""
	Parse argument options and do corresponding actions.
	"""
	# Display the banner on terminal
	Parser = argparse.ArgumentParser(description='Manage architecture and HDL libraries.', epilog="If no arguments are specified, the GUI will popup.")
	
	#------------------General command arguments-------------------------------------
	# ARGUMENT: Launch the GUI
	Parser.add_argument('-g', '--gui', action='store_true', help='If the GUI should be open (ignoring specified options).', default=False)
	# ARGUMENT: Display version
	Parser.add_argument('--version', action='version', version="%(prog)s: version='{0}'".format(VERSION), help="Displays YANGO version.")
	
	#================================================================
	SubParsers = Parser.add_subparsers(help='Library manager sub-commands.', dest='Sub-command')
	SubParsers.required = True
	EvalParser = SubParsers.add_parser('eval', help='Evaluate the resources used by a library module.')
	# Create the parser for the 'gui' command
	GUIParser = SubParsers.add_parser('gui', help='Launch the GUI.')
	# Create the parser for the 'ListModules' command
	ListModParser = SubParsers.add_parser('ListModules', help='Display a list of modules found in the library.')
	NoCModelParser= SubParsers.add_parser('NoCModel', help='Estimate NoC resource and generate NoC model. Two files are created a *.ini and *.xls.')
	#================================================================
	#------------------'GUI' command arguments---------------------------------------
	# None
	#------------------'ListModules' command arguments---------------------------------------
	# None
	#------------------'eval' command arguments---------------------------------------
	EvalParser.add_argument('-m', '--module', metavar='ModuleName', type=str, help='The name of a module saved in library.', required=True)
	EvalParser.add_argument('-a', '--architecture', metavar='ArchiName', type=str, help='The name of an Architecture saved in library.', required=True)
	EvalParser.add_argument('-p', '--parameter', metavar='ParamValue', action='append', help='A coma sparated liste of Parameter Name and value pairs (Param:Value).', default=[])
	# ARGUMENT: Display version
	EvalParser.add_argument('--host', action='store', help='Execute commands remotely.', default=None)
	#------------------'NoCModel' command arguments---------------------------------------
	NoCModelParser.add_argument('-a', '--architecture', metavar='ArchiName', type=str, help='The name of an Architecture saved in library.', required=True)
	NoCModelParser.add_argument('-o', '--output', action='store', help='Output folder path.', type=DirectoryPath, default=None)
	NoCModelParser.add_argument('--host', action='store', help='Execute commands remotely.', default=None)
	
	GUIParser.set_defaults(     func=ShowGUI)
	ListModParser.set_defaults( func=ListModules)
	EvalParser.set_defaults(    func=EvalModuleCmd)
	NoCModelParser.set_defaults(func=ModelNoCResourcesCmd)
	
	Opt = Parser.parse_args()
	Opt.func(Opt)
	return Opt
	
#====================================================================	
def DirectoryPath(Path):
	"""
	raise error if path is no a directory
	"""
	if os.path.isdir(Path):
		return os.path.abspath(Path)
	else:
		try: os.makedirs(Path)
		except:	raise TypeError
		
#====================================================================	
def FilePath(Path):
	"""
	raise error if path is no a directory
	"""
	if os.path.isfile(Path):
		return os.path.abspath(Path)
	else:
		raise TypeError
		
#====================================================================	
	
	
# ====================  START OF THE SysGen APPLICATION  =====================
if (__name__ == "__main__"):
	ColoredLogging.SetActive(True)
	# It's a standalone module
	Options = ParseOptions()
	
	logging.shutdown()
	sys.exit(0)
	
	
	
	
	
	
